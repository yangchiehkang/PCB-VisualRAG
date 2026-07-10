from pathlib import Path
from collections import defaultdict
import csv
import json
import math
import time


PROJECT_ROOT = Path(r"E:\Working\PCB_VisualRAG_Project")

QRELS_PATH = PROJECT_ROOT / "data" / "metadata" / "qrels.tsv"

C2F_RESULT_DIR = PROJECT_ROOT / "results" / "budgeted" / "coarse_to_fine"
SUMMARY_DIR = C2F_RESULT_DIR / "summary"

DAY3_RERANK_SUMMARY_PATH = C2F_RESULT_DIR / "c2f_single_vector_day3_rerank_summary.csv"
DAY2_COARSE_RECALL_PATH = C2F_RESULT_DIR / "single_vector_coarse_recall.csv"

N_VALUES = [10, 20, 50, 100]

FULL_MV_RUN_CANDIDATES = [
    PROJECT_ROOT / "results" / "full_multivector" / "full_multivector_run.tsv",
    PROJECT_ROOT / "results" / "full_multivector" / "full_mv_run.tsv",
    PROJECT_ROOT / "results" / "multivector" / "full_multivector_run.tsv",
    PROJECT_ROOT / "results" / "multivector" / "full_mv_run.tsv",
    PROJECT_ROOT / "results" / "baselines" / "full_multivector_run.tsv",
    PROJECT_ROOT / "results" / "baselines" / "full_mv_run.tsv",
    PROJECT_ROOT / "results" / "week3" / "full_multivector_run.tsv",
    PROJECT_ROOT / "results" / "week3" / "full_mv_run.tsv",
]

OUT_METRICS_CSV = SUMMARY_DIR / "c2f_day4_metrics_by_N.csv"
OUT_LATENCY_CSV = SUMMARY_DIR / "c2f_day4_latency_by_N.csv"
OUT_SUMMARY_CSV = SUMMARY_DIR / "c2f_summary.csv"
OUT_SUMMARY_XLSX = SUMMARY_DIR / "c2f_summary.xlsx"
OUT_JSON = SUMMARY_DIR / "c2f_day4_evaluation_summary.json"
OUT_MD = SUMMARY_DIR / "c2f_day4_initial_effect_latency_table.md"


def ensure_dirs():
    SUMMARY_DIR.mkdir(parents=True, exist_ok=True)


def read_tsv_rows(path):
    if not path.exists():
        raise FileNotFoundError("Missing file: {}".format(path))

    rows = []

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter="\t")
        for row in reader:
            if row and any(cell.strip() for cell in row):
                rows.append(row)

    return rows


def parse_qrels(path):
    rows = read_tsv_rows(path)

    first = [x.strip() for x in rows[0]]
    lower = [x.lower() for x in first]

    has_header = (
        ("query_id" in lower or "qid" in lower)
        and ("page_id" in lower or "doc_id" in lower)
    )

    qrels = defaultdict(dict)

    if has_header:
        header = lower
        data_rows = rows[1:]

        def find_col(candidates):
            for c in candidates:
                if c in header:
                    return header.index(c)
            return None

        q_idx = find_col(["query_id", "qid"])
        p_idx = find_col(["page_id", "doc_id"])
        rel_idx = find_col(["relevance", "rel", "label"])

        if q_idx is None or p_idx is None:
            raise ValueError("Cannot identify query/page columns in qrels header: {}".format(first))

        for row in data_rows:
            if len(row) <= max(q_idx, p_idx):
                continue

            qid = row[q_idx].strip()
            pid = row[p_idx].strip()

            rel = 1.0
            if rel_idx is not None and len(row) > rel_idx:
                try:
                    rel = float(row[rel_idx])
                except Exception:
                    rel = 1.0

            if rel > 0:
                qrels[qid][pid] = rel

    else:
        for row in rows:
            row = [x.strip() for x in row]

            if len(row) >= 4 and row[1] in ["0", "Q0"]:
                qid = row[0]
                pid = row[2]
                rel_raw = row[3]
            elif len(row) >= 3:
                qid = row[0]
                pid = row[1]
                rel_raw = row[2]
            elif len(row) >= 2:
                qid = row[0]
                pid = row[1]
                rel_raw = "1"
            else:
                continue

            try:
                rel = float(rel_raw)
            except Exception:
                rel = 1.0

            if rel > 0:
                qrels[qid][pid] = rel

    return qrels


def parse_run_file(path):
    rows = read_tsv_rows(path)

    first = [x.strip() for x in rows[0]]
    lower = [x.lower() for x in first]

    has_header = (
        ("query_id" in lower or "qid" in lower)
        and ("page_id" in lower or "doc_id" in lower)
    )

    by_query = defaultdict(list)

    if has_header:
        header = lower
        data_rows = rows[1:]

        def find_col(candidates):
            for c in candidates:
                if c in header:
                    return header.index(c)
            return None

        q_idx = find_col(["query_id", "qid"])
        p_idx = find_col(["page_id", "doc_id"])
        r_idx = find_col(["rank"])
        s_idx = find_col(["score"])

        if q_idx is None or p_idx is None:
            raise ValueError("Cannot identify query/page columns in run header: {}".format(first))

        for row in data_rows:
            if len(row) <= max(q_idx, p_idx):
                continue

            qid = row[q_idx].strip()
            pid = row[p_idx].strip()

            rank = None
            if r_idx is not None and len(row) > r_idx:
                try:
                    rank = int(float(row[r_idx]))
                except Exception:
                    rank = None

            score = 0.0
            if s_idx is not None and len(row) > s_idx:
                try:
                    score = float(row[s_idx])
                except Exception:
                    score = 0.0

            by_query[qid].append(
                {
                    "page_id": pid,
                    "rank": rank,
                    "score": score,
                }
            )

    else:
        for row in rows:
            row = [x.strip() for x in row]

            if len(row) >= 6 and row[1].upper() == "Q0":
                qid = row[0]
                pid = row[2]
                rank = int(float(row[3]))
                score = float(row[4])
            elif len(row) >= 5:
                qid = row[1]
                pid = row[2]
                rank = int(float(row[3]))
                score = float(row[4])
            elif len(row) >= 4:
                qid = row[0]
                pid = row[1]
                rank = int(float(row[2]))
                score = float(row[3])
            else:
                continue

            by_query[qid].append(
                {
                    "page_id": pid,
                    "rank": rank,
                    "score": score,
                }
            )

    for qid in by_query:
        by_query[qid].sort(
            key=lambda x: (
                x["rank"] if x["rank"] is not None else 10**9,
                -x["score"],
            )
        )

        for idx, item in enumerate(by_query[qid], start=1):
            item["rank"] = idx

    return by_query


def dcg_at_k(relevance_list, k):
    score = 0.0

    for idx, rel in enumerate(relevance_list[:k], start=1):
        score += rel / math.log2(idx + 1)

    return score


def evaluate_run(run_by_query, qrels, k_values=[1, 5, 10]):
    query_ids = sorted(qrels.keys())
    total_queries = 0

    recall_sums = {k: 0.0 for k in k_values}
    mrr10_sum = 0.0
    ndcg10_sum = 0.0

    per_query_rows = []

    for qid in query_ids:
        rel_dict = qrels[qid]
        rel_pages = set(rel_dict.keys())

        if not rel_pages:
            continue

        total_queries += 1

        ranking = run_by_query.get(qid, [])
        ranked_pages = [item["page_id"] for item in ranking]

        row = {"query_id": qid}

        for k in k_values:
            topk = ranked_pages[:k]
            hit_count = len(set(topk) & rel_pages)
            recall_k = hit_count / len(rel_pages) if rel_pages else 0.0
            recall_sums[k] += recall_k
            row["Recall@{}".format(k)] = recall_k

        rr = 0.0
        for idx, pid in enumerate(ranked_pages[:10], start=1):
            if pid in rel_pages:
                rr = 1.0 / idx
                break

        mrr10_sum += rr
        row["MRR@10"] = rr

        gains = []
        for pid in ranked_pages[:10]:
            gains.append(rel_dict.get(pid, 0.0))

        dcg = dcg_at_k(gains, 10)

        ideal_gains = sorted(rel_dict.values(), reverse=True)
        ideal_gains = ideal_gains[:10]
        idcg = dcg_at_k(ideal_gains, 10)

        ndcg = dcg / idcg if idcg > 0 else 0.0
        ndcg10_sum += ndcg
        row["nDCG@10"] = ndcg

        first_hit_rank = ""
        for idx, pid in enumerate(ranked_pages, start=1):
            if pid in rel_pages:
                first_hit_rank = idx
                break

        row["first_hit_rank"] = first_hit_rank
        row["relevant_pages"] = ";".join(sorted(rel_pages))
        row["top10_pages"] = ";".join(ranked_pages[:10])

        per_query_rows.append(row)

    metrics = {
        "evaluated_queries": total_queries,
        "Recall@1": recall_sums[1] / total_queries if total_queries else 0.0,
        "Recall@5": recall_sums[5] / total_queries if total_queries else 0.0,
        "Recall@10": recall_sums[10] / total_queries if total_queries else 0.0,
        "MRR@10": mrr10_sum / total_queries if total_queries else 0.0,
        "nDCG@10": ndcg10_sum / total_queries if total_queries else 0.0,
    }

    return metrics, per_query_rows


def find_full_mv_run():
    for path in FULL_MV_RUN_CANDIDATES:
        if path.exists():
            return path

    return None


def read_day3_latency_summary(path):
    result = {}

    if not path.exists():
        return result

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)

        for row in reader:
            try:
                n = int(float(row["N"]))
            except Exception:
                continue

            def get_float(name, default=0.0):
                try:
                    return float(row.get(name, default))
                except Exception:
                    return default

            result[n] = {
                "query_count": int(float(row.get("query_count", 0))),
                "total_candidates": int(float(row.get("total_candidates", 0))),
                "total_scored": int(float(row.get("total_scored", 0))),
                "avg_candidates_per_query": get_float("avg_candidates_per_query"),
                "avg_scored_per_query": get_float("avg_scored_per_query"),
                "rerank_time_seconds": get_float("rerank_time_seconds"),
                "avg_rerank_latency_ms_per_query": get_float("avg_rerank_latency_ms_per_query"),
            }

    return result


def read_day2_coarse_recall(path):
    result = {}

    if not path.exists():
        return result

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)

        for row in reader:
            try:
                n = int(float(row["N"]))
            except Exception:
                continue

            try:
                recall = float(row.get("coarse_recall", 0.0))
            except Exception:
                recall = 0.0

            result[n] = {
                "coarse_recall": recall,
                "hit_queries": int(float(row.get("hit_queries", 0))),
                "miss_queries": int(float(row.get("miss_queries", 0))),
                "evaluated_queries": int(float(row.get("evaluated_queries", 0))),
            }

    return result


def write_metrics_csv(rows):
    with OUT_METRICS_CSV.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "method",
                "N",
                "evaluated_queries",
                "Recall@1",
                "Recall@5",
                "Recall@10",
                "MRR@10",
                "nDCG@10",
                "run_file",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)


def write_latency_csv(rows):
    with OUT_LATENCY_CSV.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "method",
                "N",
                "query_count",
                "total_candidates",
                "total_scored",
                "avg_candidates_per_query",
                "avg_scored_per_query",
                "coarse_time_seconds_proxy",
                "rerank_time_seconds",
                "total_latency_seconds_proxy",
                "per_query_latency_ms_proxy",
                "note",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)


def write_summary_csv(rows):
    with OUT_SUMMARY_CSV.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "Method",
                "Recall@1",
                "Recall@5",
                "Recall@10",
                "MRR@10",
                "nDCG@10",
                "Latency",
                "Actual Candidates / Query",
                "Note",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(summary_rows, metrics_rows, latency_rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except Exception:
        return False

    wb = Workbook()

    ws = wb.active
    ws.title = "summary"

    summary_headers = [
        "Method",
        "Recall@1",
        "Recall@5",
        "Recall@10",
        "MRR@10",
        "nDCG@10",
        "Latency",
        "Actual Candidates / Query",
        "Note",
    ]

    ws.append(summary_headers)

    for row in summary_rows:
        ws.append([row.get(h, "") for h in summary_headers])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws2 = wb.create_sheet("metrics_by_N")

    metrics_headers = [
        "method",
        "N",
        "evaluated_queries",
        "Recall@1",
        "Recall@5",
        "Recall@10",
        "MRR@10",
        "nDCG@10",
        "run_file",
    ]

    ws2.append(metrics_headers)

    for row in metrics_rows:
        ws2.append([row.get(h, "") for h in metrics_headers])

    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws3 = wb.create_sheet("latency_by_N")

    latency_headers = [
        "method",
        "N",
        "query_count",
        "total_candidates",
        "total_scored",
        "avg_candidates_per_query",
        "avg_scored_per_query",
        "coarse_time_seconds_proxy",
        "rerank_time_seconds",
        "total_latency_seconds_proxy",
        "per_query_latency_ms_proxy",
        "note",
    ]

    ws3.append(latency_headers)

    for row in latency_rows:
        ws3.append([row.get(h, "") for h in latency_headers])

    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for sheet in [ws, ws2, ws3]:
        for column_cells in sheet.columns:
            max_len = 0
            col_letter = column_cells[0].column_letter

            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))

            sheet.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(OUT_SUMMARY_XLSX)
    return True


def write_md(summary_rows):
    lines = []

    lines.append("# Week 4 Day 4 Initial Effect-Latency Table")
    lines.append("")
    lines.append("**Project:** PCB_VisualRAG_Project  ")
    lines.append("**Stage:** Week 4 Day 4  ")
    lines.append("**Experiment:** Coarse-to-Fine Evaluation  ")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## Table 4: Full MV 与 Coarse-to-Fine 检索结果对比")
    lines.append("")
    lines.append("| Method | Recall@10 | MRR@10 | nDCG@10 | Latency | Actual Candidates / Query | Note |")
    lines.append("|---|---:|---:|---:|---:|---:|---|")

    for row in summary_rows:
        lines.append(
            "| {} | {} | {} | {} | {} | {} | {} |".format(
                row["Method"],
                row["Recall@10"],
                row["MRR@10"],
                row["nDCG@10"],
                row["Latency"],
                row["Actual Candidates / Query"],
                row["Note"],
            )
        )

    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## Key Observation")
    lines.append("")
    lines.append("当前 Day 4 的评测结果需要结合 Day 2 和 Day 3 一起解读。")
    lines.append("")
    lines.append("- Day 2 显示 single-vector coarse recall 较低；")
    lines.append("- Day 3 显示 N=10、20、50、100 的实际候选数均为每 query 10 个；")
    lines.append("- 因此 Day 4 中不同 N 的效果与时延差异可能不明显；")
    lines.append("- 当前实验主要验证 C2F pipeline 的完整闭环，而不是最终的候选预算曲线。")
    lines.append("")

    OUT_MD.write_text("\n".join(lines), encoding="utf-8")


def format_float(x):
    if x == "":
        return ""
    return "{:.4f}".format(float(x))


def main():
    ensure_dirs()

    print("[Day4] Loading qrels...")
    qrels = parse_qrels(QRELS_PATH)
    print("[Day4] Loaded qrels queries: {}".format(len(qrels)))

    day3_latency = read_day3_latency_summary(DAY3_RERANK_SUMMARY_PATH)
    day2_coarse = read_day2_coarse_recall(DAY2_COARSE_RECALL_PATH)

    metrics_rows = []
    latency_rows = []
    summary_rows = []

    full_mv_run_path = find_full_mv_run()

    if full_mv_run_path is not None:
        print("[Day4] Found Full MV run: {}".format(full_mv_run_path))

        start = time.perf_counter()
        full_mv_run = parse_run_file(full_mv_run_path)
        metrics, _ = evaluate_run(full_mv_run, qrels)
        elapsed = time.perf_counter() - start

        metrics_rows.append(
            {
                "method": "Full Multi-vector",
                "N": "full",
                "evaluated_queries": metrics["evaluated_queries"],
                "Recall@1": "{:.4f}".format(metrics["Recall@1"]),
                "Recall@5": "{:.4f}".format(metrics["Recall@5"]),
                "Recall@10": "{:.4f}".format(metrics["Recall@10"]),
                "MRR@10": "{:.4f}".format(metrics["MRR@10"]),
                "nDCG@10": "{:.4f}".format(metrics["nDCG@10"]),
                "run_file": str(full_mv_run_path.relative_to(PROJECT_ROOT)),
            }
        )

        summary_rows.append(
            {
                "Method": "Full Multi-vector",
                "Recall@1": "{:.4f}".format(metrics["Recall@1"]),
                "Recall@5": "{:.4f}".format(metrics["Recall@5"]),
                "Recall@10": "{:.4f}".format(metrics["Recall@10"]),
                "MRR@10": "{:.4f}".format(metrics["MRR@10"]),
                "nDCG@10": "{:.4f}".format(metrics["nDCG@10"]),
                "Latency": "",
                "Actual Candidates / Query": "full corpus",
                "Note": "Full MV run found; latency should be filled from Week 3 cost log if available.",
            }
        )
    else:
        print("[Day4] Full MV run not found. Full MV row will be left blank.")

        summary_rows.append(
            {
                "Method": "Full Multi-vector",
                "Recall@1": "",
                "Recall@5": "",
                "Recall@10": "",
                "MRR@10": "",
                "nDCG@10": "",
                "Latency": "",
                "Actual Candidates / Query": "full corpus",
                "Note": "Full MV run file not auto-detected. Fill from Week 3 result if needed.",
            }
        )

    for n in N_VALUES:
        run_path = C2F_RESULT_DIR / "c2f_single_vector_N{}_run.tsv".format(n)

        if not run_path.exists():
            alt_path = C2F_RESULT_DIR / "c2f_N{}_run.tsv".format(n)
            if alt_path.exists():
                run_path = alt_path

        if not run_path.exists():
            print("[Day4] Missing C2F run for N={}: skipped".format(n))
            continue

        print("[Day4] Evaluating C2F N={} ...".format(n))

        run_by_query = parse_run_file(run_path)
        metrics, per_query_rows = evaluate_run(run_by_query, qrels)

        per_query_path = SUMMARY_DIR / "c2f_N{}_per_query_metrics.csv".format(n)

        with per_query_path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=[
                    "query_id",
                    "Recall@1",
                    "Recall@5",
                    "Recall@10",
                    "MRR@10",
                    "nDCG@10",
                    "first_hit_rank",
                    "relevant_pages",
                    "top10_pages",
                ],
            )
            writer.writeheader()
            writer.writerows(per_query_rows)

        latency = day3_latency.get(n, {})
        coarse = day2_coarse.get(n, {})

        query_count = latency.get("query_count", metrics["evaluated_queries"])
        total_candidates = latency.get("total_candidates", "")
        total_scored = latency.get("total_scored", "")
        avg_candidates_per_query = latency.get("avg_candidates_per_query", "")
        avg_scored_per_query = latency.get("avg_scored_per_query", "")

        rerank_time_seconds = latency.get("rerank_time_seconds", 0.0)

        coarse_time_seconds_proxy = 0.0
        total_latency_seconds_proxy = coarse_time_seconds_proxy + rerank_time_seconds

        if query_count:
            per_query_latency_ms_proxy = total_latency_seconds_proxy * 1000.0 / query_count
        else:
            per_query_latency_ms_proxy = 0.0

        metrics_rows.append(
            {
                "method": "C2F N={}".format(n),
                "N": n,
                "evaluated_queries": metrics["evaluated_queries"],
                "Recall@1": "{:.4f}".format(metrics["Recall@1"]),
                "Recall@5": "{:.4f}".format(metrics["Recall@5"]),
                "Recall@10": "{:.4f}".format(metrics["Recall@10"]),
                "MRR@10": "{:.4f}".format(metrics["MRR@10"]),
                "nDCG@10": "{:.4f}".format(metrics["nDCG@10"]),
                "run_file": str(run_path.relative_to(PROJECT_ROOT)),
            }
        )

        latency_note = (
            "coarse_time_seconds_proxy=0 because Day 2 used cached coarse candidates; "
            "rerank_time_seconds comes from Day 3 summary."
        )

        if avg_candidates_per_query == 10.0 and n != 10:
            latency_note += " Actual candidate depth is still 10, not {}.".format(n)

        latency_rows.append(
            {
                "method": "C2F N={}".format(n),
                "N": n,
                "query_count": query_count,
                "total_candidates": total_candidates,
                "total_scored": total_scored,
                "avg_candidates_per_query": "{:.4f}".format(avg_candidates_per_query) if avg_candidates_per_query != "" else "",
                "avg_scored_per_query": "{:.4f}".format(avg_scored_per_query) if avg_scored_per_query != "" else "",
                "coarse_time_seconds_proxy": "{:.6f}".format(coarse_time_seconds_proxy),
                "rerank_time_seconds": "{:.6f}".format(rerank_time_seconds),
                "total_latency_seconds_proxy": "{:.6f}".format(total_latency_seconds_proxy),
                "per_query_latency_ms_proxy": "{:.6f}".format(per_query_latency_ms_proxy),
                "note": latency_note,
            }
        )

        coarse_recall_text = ""
        if coarse:
            coarse_recall_text = "coarse recall={:.4f}; ".format(coarse.get("coarse_recall", 0.0))

        note = coarse_recall_text
        if avg_candidates_per_query == 10.0 and n != 10:
            note += "actual candidates/query is 10 due to limited coarse run depth."
        elif avg_candidates_per_query != "":
            note += "actual candidates/query={:.4f}.".format(avg_candidates_per_query)
        else:
            note += "latency summary unavailable."

        summary_rows.append(
            {
                "Method": "C2F N={}".format(n),
                "Recall@1": "{:.4f}".format(metrics["Recall@1"]),
                "Recall@5": "{:.4f}".format(metrics["Recall@5"]),
                "Recall@10": "{:.4f}".format(metrics["Recall@10"]),
                "MRR@10": "{:.4f}".format(metrics["MRR@10"]),
                "nDCG@10": "{:.4f}".format(metrics["nDCG@10"]),
                "Latency": "{:.6f} ms/query".format(per_query_latency_ms_proxy),
                "Actual Candidates / Query": "{:.4f}".format(avg_candidates_per_query) if avg_candidates_per_query != "" else "",
                "Note": note,
            }
        )

    write_metrics_csv(metrics_rows)
    write_latency_csv(latency_rows)
    write_summary_csv(summary_rows)

    xlsx_ok = write_xlsx(summary_rows, metrics_rows, latency_rows)
    write_md(summary_rows)

    payload = {
        "metrics_rows": metrics_rows,
        "latency_rows": latency_rows,
        "summary_rows": summary_rows,
        "outputs": {
            "metrics_csv": str(OUT_METRICS_CSV.relative_to(PROJECT_ROOT)),
            "latency_csv": str(OUT_LATENCY_CSV.relative_to(PROJECT_ROOT)),
            "summary_csv": str(OUT_SUMMARY_CSV.relative_to(PROJECT_ROOT)),
            "summary_xlsx": str(OUT_SUMMARY_XLSX.relative_to(PROJECT_ROOT)) if xlsx_ok else "",
            "summary_md": str(OUT_MD.relative_to(PROJECT_ROOT)),
        },
    }

    with OUT_JSON.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    print("")
    print("========== Day 4 Completed ==========")
    print("Metrics CSV:  {}".format(OUT_METRICS_CSV))
    print("Latency CSV:  {}".format(OUT_LATENCY_CSV))
    print("Summary CSV:  {}".format(OUT_SUMMARY_CSV))
    print("Summary MD:   {}".format(OUT_MD))

    if xlsx_ok:
        print("Summary XLSX: {}".format(OUT_SUMMARY_XLSX))
    else:
        print("Summary XLSX: not generated because openpyxl is not installed.")
        print("Install command: pip install openpyxl")
        print("Then rerun: python scripts\\evaluation\\evaluate_c2f_day4.py")

    print("")
    print("Table 4 Preview:")
    print("Method, Recall@10, MRR@10, nDCG@10, Latency, Actual Candidates / Query")

    for row in summary_rows:
        print(
            "{}, {}, {}, {}, {}, {}".format(
                row["Method"],
                row["Recall@10"],
                row["MRR@10"],
                row["nDCG@10"],
                row["Latency"],
                row["Actual Candidates / Query"],
            )
        )


if __name__ == "__main__":
    main()
